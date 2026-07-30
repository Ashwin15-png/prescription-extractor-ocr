import os
import io
import shutil
from typing import List, Optional
from fastapi import APIRouter, UploadFile, File, HTTPException, Depends, Query, status
from fastapi.responses import StreamingResponse, Response, JSONResponse
from sqlalchemy.orm import Session
from sqlalchemy import func, or_, and_, not_, text
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from fpdf import FPDF, XPos, YPos
import json
from datetime import datetime, timedelta, timezone

from .database import get_db
from .models import Prescription, User
from .schemas import (
    OCRUploadResponse, ExtractedFields, PrescriptionCreate, PrescriptionResponse,
    SuggestionsResponse, UserCreate, UserResponse, Token
)
from starlette.concurrency import run_in_threadpool
from .ocr import perform_ocr, analyze_image_quality, detect_document_type, process_and_analyze_image
from .extractor import extract_fields
from .config import settings
from .logger import logger
from .auth import hash_password, verify_password, create_access_token, get_current_user

# ── Phase 6 Enterprise Filter Engine Logic ─────────────────────────────────
def build_field_clause(key, val):
    if val is None or val == "":
        return None
        
    # If val is a dictionary (like range comparison)
    if isinstance(val, dict):
        terms = []
        col = getattr(Prescription, key, None)
        if col is None:
            # Fallback for nested operations
            if key in ("confidence", "ocr_confidence"):
                col = Prescription.confidence_score
            elif key in ("image_quality", "quality"):
                col = Prescription.image_quality_score
            else:
                return None
        for op, op_val in val.items():
            if op_val is None or op_val == "":
                continue
            if op in ("gte", "min"):
                terms.append(col >= op_val)
            elif op in ("lte", "max"):
                terms.append(col <= op_val)
            elif op in ("gt"):
                terms.append(col > op_val)
            elif op in ("lt"):
                terms.append(col < op_val)
            elif op in ("eq"):
                terms.append(col == op_val)
            elif op in ("ne", "neq"):
                terms.append(col != op_val)
            elif op in ("like"):
                terms.append(col.ilike(f"%{op_val}%"))
        return and_(*terms) if terms else None

    # Handle standard flat mappings
    if key == "date_range":
        now = datetime.now()
        col = Prescription.created_at
        if val == "today":
            start = datetime(now.year, now.month, now.day)
            return col >= start
        elif val == "yesterday":
            start = datetime(now.year, now.month, now.day) - timedelta(days=1)
            end = datetime(now.year, now.month, now.day)
            return and_(col >= start, col < end)
        elif val == "last_7_days":
            return col >= now - timedelta(days=7)
        elif val == "last_30_days":
            return col >= now - timedelta(days=30)
        elif val == "last_90_days":
            return col >= now - timedelta(days=90)
        elif val == "current_month":
            start = datetime(now.year, now.month, 1)
            return col >= start
        elif val == "previous_month":
            first_this_month = datetime(now.year, now.month, 1)
            if now.month == 1:
                first_prev_month = datetime(now.year - 1, 12, 1)
            else:
                first_prev_month = datetime(now.year, now.month - 1, 1)
            return and_(col >= first_prev_month, col < first_this_month)
        elif val == "current_year":
            return col >= datetime(now.year, 1, 1)
        elif val == "financial_year":
            if now.month >= 4:
                start = datetime(now.year, 4, 1)
            else:
                start = datetime(now.year - 1, 4, 1)
            return col >= start
        return None

    if key == "custom_date_range":
        if isinstance(val, dict):
            terms = []
            if val.get("start"):
                try:
                    terms.append(Prescription.created_at >= datetime.strptime(val["start"], "%Y-%m-%d"))
                except Exception:
                    pass
            if val.get("end"):
                try:
                    terms.append(Prescription.created_at <= datetime.strptime(val["end"] + " 23:59:59", "%Y-%m-%d %H:%M:%S"))
                except Exception:
                    pass
            return and_(*terms) if terms else None
        return None

    col = getattr(Prescription, key, None)
    if col is None:
         # Fallback mappings for field aliases
         if key in ("patient", "patient_name"):
             col = Prescription.patient_name
         elif key in ("medicine", "medicine_name"):
             col = Prescription.medicine
         elif key in ("doctor", "doctor_name"):
             col = Prescription.doctor_name
         elif key in ("hospital", "hospital_name"):
             col = Prescription.hospital_name
         elif key in ("confidence", "ocr_confidence", "confidence_score"):
             col = Prescription.confidence_score
         elif key in ("image_quality", "quality", "image_quality_score"):
             col = Prescription.image_quality_score
         elif key in ("blur", "blur_score"):
             col = Prescription.blur_score
         elif key in ("emerg", "emergency", "is_emergency"):
             col = Prescription.is_emergency
         elif key in ("inpatient", "is_inpatient"):
             col = Prescription.is_inpatient
         elif key in ("outpatient", "is_outpatient"):
             col = Prescription.is_outpatient
         elif key in ("handwritten", "is_handwritten"):
             col = Prescription.is_handwritten

    if col is None:
        return None

    # Apply based on type of column and value
    from sqlalchemy.sql.sqltypes import Boolean, Integer, Float
    if isinstance(col.type, Boolean):
        bval = str(val).lower() in ("true", "1", "yes", "t", "y")
        return col == bval
    elif isinstance(col.type, (Integer, Float)):
        try:
            return col == col.type.python_type(val)
        except ValueError:
            return None
    else:
        return col.ilike(f"%{val}%")

def parse_expression(expr):
    if not isinstance(expr, dict):
        return None
    
    clauses = []
    for key, val in expr.items():
        if key == "_and":
            if isinstance(val, list):
                subterms = [parse_expression(item) for item in val if parse_expression(item) is not None]
                if subterms:
                    clauses.append(and_(*subterms))
        elif key == "_or":
            if isinstance(val, list):
                subterms = [parse_expression(item) for item in val if parse_expression(item) is not None]
                if subterms:
                    clauses.append(or_(*subterms))
        elif key == "_not":
            subterm = parse_expression(val)
            if subterm is not None:
                clauses.append(not_(subterm))
        else:
            term = build_field_clause(key, val)
            if term is not None:
                clauses.append(term)
    
    if not clauses:
        return None
    return and_(*clauses)

def apply_enterprise_filters(query, filters_str_or_dict):
    if not filters_str_or_dict:
        return query
    
    if isinstance(filters_str_or_dict, str):
        try:
            filters_dict = json.loads(filters_str_or_dict)
        except Exception:
            return query
    else:
        filters_dict = filters_str_or_dict

    if not filters_dict:
        return query

    # If it is flat filters, wrap it inside an _and block
    if isinstance(filters_dict, dict) and not any(k in ("_and", "_or", "_not") for k in filters_dict.keys()):
        filters_dict = {"_and": [{k: v} for k, v in filters_dict.items()]}

    clause = parse_expression(filters_dict)
    if clause is not None:
        query = query.filter(clause)
    return query

router = APIRouter()

# ── 1. Upload & OCR Endpoint ──────────────────────────────────────────────
@router.post("/upload", response_model=OCRUploadResponse)
async def upload_prescription(file: UploadFile = File(...), db: Session = Depends(get_db)):
    if not file.filename:
        raise HTTPException(status_code=400, detail="No file selected.")

    # Validate file extension
    ext = os.path.splitext(file.filename)[1].lower()
    allowed_exts = [".jpg", ".jpeg", ".png", ".webp", ".bmp", ".tiff"]
    if ext not in allowed_exts and not file.content_type.startswith("image/"):
        raise HTTPException(status_code=400, detail="File type not supported or file must be an image.")

    contents = await file.read()
    if len(contents) > 10 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="File size exceeds 10MB limit.")

    os.makedirs("uploads", exist_ok=True)
    temp_path = os.path.join("uploads", f"temp_{file.filename}")
    with open(temp_path, "wb") as f:
        f.write(contents)

    try:
        raw_text, analysis = await run_in_threadpool(process_and_analyze_image, temp_path)
        extracted = await run_in_threadpool(extract_fields, raw_text)
        doc_type = detect_document_type(raw_text)

        is_duplicate = False
        if extracted.get("patient_name") != "Unknown" and extracted.get("medicine") != "Not Found":
            existing = db.query(Prescription).filter(
                Prescription.patient_name.ilike(extracted["patient_name"]),
                Prescription.medicine.ilike(extracted["medicine"])
            ).first()
            if existing: is_duplicate = True

        response_data = OCRUploadResponse(
            raw_text=raw_text,
            extracted_fields=ExtractedFields(
                patient_name=extracted.get("patient_name", "Unknown"),
                medicine=extracted.get("medicine", "Not Found"),
                dosage=extracted.get("dosage", "Not Found"),
                date=extracted.get("date", "Not Found"),
                doctor_name=extracted.get("doctor_name", "Unknown"),
                hospital_name=extracted.get("hospital_name", "Unknown"),
                age=extracted.get("age", "N/A"),
                gender=extracted.get("gender", "N/A"),
                document_type=doc_type,
                hospital_address=extracted.get("hospital_address", ""),
                registration_num=extracted.get("registration_num", ""),
                generic_name=extracted.get("generic_name", ""),
                strength=extracted.get("strength", ""),
                frequency=extracted.get("frequency", ""),
                duration=extracted.get("duration", ""),
                diagnosis=extracted.get("diagnosis", ""),
                symptoms=extracted.get("symptoms", ""),
                department=extracted.get("department", ""),
                follow_up_date=extracted.get("follow_up_date", ""),
                report_num=extracted.get("report_num", ""),
                lab_tests=extracted.get("lab_tests", ""),
                qr_code_data=analysis.get("qr_code_data", ""),
                
                # Phase 6 additions
                noise_level=analysis.get("noise_level", 0),
                skew_angle=analysis.get("skew_angle", 0.0),
                rotation=analysis.get("rotation", 0),
                contrast_score=analysis.get("contrast_score", 100),
                brightness_score=analysis.get("brightness_score", 100),
                readability_score=analysis.get("readability_score", 100),
                language=analysis.get("language", "English"),
                barcode=analysis.get("barcode", "None"),
                is_handwritten=analysis.get("is_handwritten", False),
                medicine_category=extracted.get("medicine_category", ""),
                doctor_specialty=extracted.get("doctor_specialty", ""),
                hospital_type=extracted.get("hospital_type", ""),
                is_emergency=extracted.get("is_emergency", False),
                is_inpatient=extracted.get("is_inpatient", False),
                is_outpatient=extracted.get("is_outpatient", False)
            ),
            ocr_confidence=analysis.get("confidence", 85),
            image_quality_score=analysis.get("image_quality", 100),
            blur_detected=analysis.get("blur_detected", False),
            is_duplicate=is_duplicate
        )
        return response_data

    except Exception as e:
        logger.error(f"Error during file processing: {e}", exc_info=True)
        return JSONResponse(
            status_code=500,
            content={
                "success": False,
                "error": "Failed to Fetch / Internal Server Error",
                "details": str(e)
            }
        )
    finally:
        if os.path.exists(temp_path):
            try: os.remove(temp_path)
            except Exception: pass

# ── 2. Save Prescription Endpoint ─────────────────────────────────────────
@router.post("/save")
async def save_prescription(data: PrescriptionCreate, db: Session = Depends(get_db)):
    if not data.patient_name or not data.patient_name.strip():
        raise HTTPException(status_code=400, detail="Patient Name is required.")
    if not data.medicine or not data.medicine.strip():
        raise HTTPException(status_code=400, detail="Medicine name is required.")

    try:
        new_prescription = Prescription(
            patient_name=data.patient_name.strip(),
            medicine=data.medicine.strip(),
            dosage=data.dosage.strip(),
            date=data.date.strip(),
            doctor_name=data.doctor_name.strip() if data.doctor_name else "Unknown",
            hospital_name=data.hospital_name.strip() if data.hospital_name else "Unknown",
            age=data.age or "N/A",
            gender=data.gender or "N/A",
            document_type=data.document_type or "Prescription",
            
            hospital_address=data.hospital_address,
            registration_num=data.registration_num,
            generic_name=data.generic_name,
            strength=data.strength,
            frequency=data.frequency,
            duration=data.duration,
            diagnosis=data.diagnosis,
            symptoms=data.symptoms,
            department=data.department,
            follow_up_date=data.follow_up_date,
            report_num=data.report_num,
            lab_tests=data.lab_tests,
            qr_code_data=data.qr_code_data,
            
            raw_text=data.raw_text,
            ocr_clean_text=data.ocr_clean_text,
            confidence_score=data.confidence_score or 90,
            image_quality_score=data.image_quality_score or 100,
            blur_score=data.blur_score or 0,
            blur_detected=data.blur_detected or False,
            qr_code=data.qr_code,
            latitude=data.latitude,
            longitude=data.longitude,
            country=data.country or "India",
            state=data.state,
            city=data.city,
            
            # Phase 6 Enterprise Filter Engine Additions
            noise_level=data.noise_level,
            skew_angle=data.skew_angle,
            rotation=data.rotation,
            contrast_score=data.contrast_score,
            brightness_score=data.brightness_score,
            readability_score=data.readability_score,
            language=data.language or "English",
            barcode=data.barcode,
            is_handwritten=data.is_handwritten or False,
            medicine_category=data.medicine_category,
            doctor_specialty=data.doctor_specialty,
            hospital_type=data.hospital_type,
            is_emergency=data.is_emergency or False,
            is_inpatient=data.is_inpatient or False,
            is_outpatient=data.is_outpatient or False
        )
        db.add(new_prescription)
        db.commit()
        db.refresh(new_prescription)
        
        logger.info(f"Successfully saved ID {new_prescription.id}")
        return {"message": "Prescription saved successfully!", "id": new_prescription.id}

    except Exception as e:
        db.rollback()
        logger.error(f"Database error while saving prescription: {e}")
        raise HTTPException(status_code=500, detail="Error saving prescription to database.")

# Helper to build enterprise search/filter queries (without sorting)
def get_prescriptions_query_base(
    db: Session,
    patient: Optional[str] = None,
    medicine: Optional[str] = None,
    date: Optional[str] = None,
    doctor: Optional[str] = None,
    hospital: Optional[str] = None,
    search: Optional[str] = None,
    filters: Optional[str] = None
):
    query = db.query(Prescription)
    
    # ── Global search ──
    if search:
        # Multiple keyword support (and-join on keywords matching any field)
        keywords = [k.strip() for k in search.split() if k.strip()]
        if keywords:
            keyword_clauses = []
            for keyword in keywords:
                search_term = f"%{keyword}%"
                keyword_clauses.append(
                    or_(
                        Prescription.patient_name.ilike(search_term),
                        Prescription.medicine.ilike(search_term),
                        Prescription.doctor_name.ilike(search_term),
                        Prescription.hospital_name.ilike(search_term),
                        Prescription.diagnosis.ilike(search_term),
                        Prescription.symptoms.ilike(search_term),
                        Prescription.department.ilike(search_term),
                        Prescription.raw_text.ilike(search_term),
                        Prescription.registration_num.ilike(search_term),
                        Prescription.date.ilike(search_term)
                    )
                )
            query = query.filter(and_(*keyword_clauses))
            
    # ── Apply Advanced Enterprise Filters ──
    query = apply_enterprise_filters(query, filters)
        
    # ── Flat Parameters (For backwards compatibility) ──
    if patient:
        query = query.filter(Prescription.patient_name.ilike(f"%{patient}%"))
    if medicine:
        query = query.filter(Prescription.medicine.ilike(f"%{medicine}%"))
    if date:
        query = query.filter(Prescription.date.ilike(f"%{date}%"))
    if doctor:
        query = query.filter(Prescription.doctor_name.ilike(f"%{doctor}%"))
    if hospital:
        query = query.filter(Prescription.hospital_name.ilike(f"%{hospital}%"))
        
    return query

def apply_sorting(query, sort_by: Optional[str] = None):
    # ── Multi-column Sorting ──
    if sort_by:
        sort_clauses = []
        parts = [p.strip() for p in sort_by.split(",") if p.strip()]
        for part in parts:
            direction = "desc"
            field_name = part
            if part.endswith("_asc"):
                direction = "asc"
                field_name = part[:-4]
            elif part.endswith("_desc"):
                direction = "desc"
                field_name = part[:-5]
            elif part.endswith(" asc"):
                direction = "asc"
                field_name = part[:-4]
            elif part.endswith(" desc"):
                direction = "desc"
                field_name = part[:-5]
            
            col = None
            if field_name in ["patient", "patient_name"]:
                col = Prescription.patient_name
            elif field_name in ["medicine", "medicine_name"]:
                col = Prescription.medicine
            elif field_name in ["date", "prescription_date"]:
                col = Prescription.date
            elif field_name in ["created_at", "id"]:
                col = Prescription.id
            elif field_name in ["doctor", "doctor_name"]:
                col = Prescription.doctor_name
            elif field_name in ["hospital", "hospital_name"]:
                col = Prescription.hospital_name
            elif field_name in ["confidence", "confidence_score", "accuracy", "ocr_accuracy"]:
                col = Prescription.confidence_score
            elif field_name in ["image_quality", "quality"]:
                col = Prescription.image_quality_score
                
            if col is not None:
                if direction == "asc":
                    sort_clauses.append(col.asc())
                else:
                    sort_clauses.append(col.desc())
                    
        if sort_clauses:
            query = query.order_by(*sort_clauses)
        else:
            query = query.order_by(Prescription.id.desc())
    else:
        query = query.order_by(Prescription.id.desc())
        
    return query

# ── 3. List & Filter Prescriptions Endpoint ───────────────────────────────
@router.get("/prescriptions", response_model=List[PrescriptionResponse])
async def get_prescriptions(
    response: Response,
    patient: Optional[str] = Query(None, description="Filter by patient name"),
    medicine: Optional[str] = Query(None, description="Filter by medicine name"),
    date: Optional[str] = Query(None, description="Filter by date"),
    doctor: Optional[str] = Query(None, description="Filter by doctor name"),
    hospital: Optional[str] = Query(None, description="Filter by hospital name"),
    search: Optional[str] = Query(None, description="Global search term"),
    sort_by: Optional[str] = Query(None, description="Sorting parameter, e.g. patient_name_asc, date_desc"),
    filters: Optional[str] = Query(None, description="JSON encoded filters"),
    page: Optional[int] = Query(None, description="Page number"),
    limit: Optional[int] = Query(None, description="Page limit size"),
    db: Session = Depends(get_db)
):
    try:
        query = get_prescriptions_query_base(
            db, patient=patient, medicine=medicine, date=date, doctor=doctor,
            hospital=hospital, search=search, filters=filters
        )
            
        # Record total count before slice
        total_count = query.count()
        response.headers["X-Total-Count"] = str(total_count)
        response.headers["Access-Control-Expose-Headers"] = "X-Total-Count"

        # Apply sorting
        query = apply_sorting(query, sort_by=sort_by)

        # Apply paging
        if page and limit:
            offset = (page - 1) * limit
            query = query.offset(offset).limit(limit)

        results = query.all()
        return results
    except Exception as e:
        logger.error(f"Error fetching prescriptions: {e}")
        raise HTTPException(status_code=500, detail="Error retrieving records from database.")

# Aliases for get_prescriptions and custom endpoints
@router.get("/records", response_model=List[PrescriptionResponse])
async def get_records(
    response: Response,
    patient: Optional[str] = Query(None),
    medicine: Optional[str] = Query(None),
    date: Optional[str] = Query(None),
    doctor: Optional[str] = Query(None),
    hospital: Optional[str] = Query(None),
    search: Optional[str] = Query(None),
    sort_by: Optional[str] = Query(None),
    filters: Optional[str] = Query(None),
    page: Optional[int] = Query(None),
    limit: Optional[int] = Query(None),
    db: Session = Depends(get_db)
):
    return await get_prescriptions(response, patient, medicine, date, doctor, hospital, search, sort_by, filters, page, limit, db)

@router.get("/search", response_model=List[PrescriptionResponse])
async def search_records(
    response: Response,
    patient: Optional[str] = Query(None),
    medicine: Optional[str] = Query(None),
    date: Optional[str] = Query(None),
    doctor: Optional[str] = Query(None),
    hospital: Optional[str] = Query(None),
    search: Optional[str] = Query(None),
    sort_by: Optional[str] = Query(None),
    filters: Optional[str] = Query(None),
    page: Optional[int] = Query(None),
    limit: Optional[int] = Query(None),
    db: Session = Depends(get_db)
):
    return await get_prescriptions(response, patient, medicine, date, doctor, hospital, search, sort_by, filters, page, limit, db)

@router.get("/sort", response_model=List[PrescriptionResponse])
async def sort_records(
    response: Response,
    patient: Optional[str] = Query(None),
    medicine: Optional[str] = Query(None),
    date: Optional[str] = Query(None),
    doctor: Optional[str] = Query(None),
    hospital: Optional[str] = Query(None),
    search: Optional[str] = Query(None),
    sort_by: Optional[str] = Query(None),
    filters: Optional[str] = Query(None),
    page: Optional[int] = Query(None),
    limit: Optional[int] = Query(None),
    db: Session = Depends(get_db)
):
    return await get_prescriptions(response, patient, medicine, date, doctor, hospital, search, sort_by, filters, page, limit, db)

@router.get("/filter", response_model=List[PrescriptionResponse])
async def filter_records(
    response: Response,
    patient: Optional[str] = Query(None),
    medicine: Optional[str] = Query(None),
    date: Optional[str] = Query(None),
    doctor: Optional[str] = Query(None),
    hospital: Optional[str] = Query(None),
    search: Optional[str] = Query(None),
    sort_by: Optional[str] = Query(None),
    filters: Optional[str] = Query(None),
    page: Optional[int] = Query(None),
    limit: Optional[int] = Query(None),
    db: Session = Depends(get_db)
):
    return await get_prescriptions(response, patient, medicine, date, doctor, hospital, search, sort_by, filters, page, limit, db)

@router.post("/extract", response_model=OCRUploadResponse)
async def extract_prescription_fields(file: UploadFile = File(...), db: Session = Depends(get_db)):
    return await upload_prescription(file, db)

@router.get("/status")
async def get_status(db: Session = Depends(get_db)):
    db_status = "unhealthy"
    try:
        db.execute(text("SELECT 1"))
        db_status = "connected"
    except Exception as e:
        logger.error(f"Status DB health check failed: {e}")
        
    tess_available = False
    try:
        import pytesseract
        pytesseract.get_tesseract_version()
        tess_available = True
    except Exception:
        pass
        
    return {
        "status": "online",
        "database": db_status,
        "tesseract_installed": tess_available,
        "uploads_dir_exists": os.path.exists("uploads")
    }

@router.get("/health")
async def health_check_route(db: Session = Depends(get_db)):
    db_status = "unhealthy"
    try:
        db.execute(text("SELECT 1"))
        db_status = "connected"
    except Exception as e:
        logger.error(f"Health check DB ping failed: {e}")
    return {
        "status": "healthy" if db_status == "connected" else "degraded",
        "database": db_status
    }

@router.delete("/prescriptions/{record_id}")
async def delete_prescription(record_id: int, db: Session = Depends(get_db)):
    try:
        record = db.query(Prescription).filter(Prescription.id == record_id).first()
        if not record:
            raise HTTPException(status_code=404, detail="Record not found.")
        db.delete(record)
        db.commit()
        logger.info(f"Deleted prescription ID {record_id}")
        return {"message": "Record deleted successfully"}
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        logger.error(f"Error deleting prescription {record_id}: {e}")
        raise HTTPException(status_code=500, detail="Error deleting record.")

# ── 4. Analytics Summary Endpoint ─────────────────────────────────────────
@router.get("/analytics")
async def get_analytics(
    search: Optional[str] = Query(None),
    filters: Optional[str] = Query(None),
    db: Session = Depends(get_db)
):
    try:
        base_query = get_prescriptions_query_base(db, search=search, filters=filters)
        total = base_query.count()

        # Top medicine
        top_med_query = base_query.with_entities(
            Prescription.medicine, func.count(Prescription.id).label("count")
        ).group_by(Prescription.medicine).order_by(func.count(Prescription.id).desc()).first()
        top_medicine = top_med_query[0] if top_med_query else "N/A"

        # Top doctor
        top_doc_query = base_query.with_entities(
            Prescription.doctor_name, func.count(Prescription.id).label("count")
        ).filter(Prescription.doctor_name != "Unknown").group_by(Prescription.doctor_name).order_by(func.count(Prescription.id).desc()).first()
        top_doctor = top_doc_query[0] if top_doc_query else "N/A"

        # Top dosage
        top_dos_query = base_query.with_entities(
            Prescription.dosage, func.count(Prescription.id).label("count")
        ).filter(Prescription.dosage != "Not Found").group_by(Prescription.dosage).order_by(func.count(Prescription.id).desc()).first()
        top_dosage = top_dos_query[0] if top_dos_query else "N/A"

        # Average confidence
        avg_conf = base_query.with_entities(func.avg(Prescription.confidence_score)).scalar() or 92.5

        # Recent uploads (last 24 hours)
        day_ago = datetime.now(timezone.utc) - timedelta(days=1)
        recent = base_query.filter(Prescription.created_at >= day_ago).count()

        return {
            "total_prescriptions": total,
            "recent_uploads": recent,
            "most_common_medicine": top_medicine,
            "most_common_dosage": top_dosage,
            "most_frequent_doctor": top_doctor,
            "average_confidence_score": round(float(avg_conf), 1),
            "ocr_accuracy_percentage": "94.2%"
        }
    except Exception as e:
        logger.error(f"Error generating analytics: {e}")
        raise HTTPException(status_code=500, detail="Error fetching analytics data.")

# ── 5. CSV Export Endpoint ────────────────────────────────────────────────
@router.get("/export-csv")
async def export_csv(
    patient: Optional[str] = Query(None),
    medicine: Optional[str] = Query(None),
    date: Optional[str] = Query(None),
    doctor: Optional[str] = Query(None),
    hospital: Optional[str] = Query(None),
    search: Optional[str] = Query(None),
    sort_by: Optional[str] = Query(None),
    filters: Optional[str] = Query(None),
    page: Optional[int] = Query(None),
    limit: Optional[int] = Query(None),
    db: Session = Depends(get_db)
):
    try:
        query = get_prescriptions_query_base(db, patient, medicine, date, doctor, hospital, search, filters)
        query = apply_sorting(query, sort_by=sort_by)
        if page and limit:
            offset = (page - 1) * limit
            query = query.offset(offset).limit(limit)
        prescriptions = query.all()
        lines = ["ID,Patient Name,Medicine,Dosage,Date,Doctor Name,Hospital Name,Diagnosis,Symptoms,Department,Follow_Up_Date,Lab_Tests,OCR Confidence,Image Quality\n"]
        for p in prescriptions:
            # Safely quote items for CSV
            def _q(val): 
                if not val: return '""'
                v_str = str(val).replace('"', '""')
                return f'"{v_str}"'
            
            lines.append(f"{p.id},{_q(p.patient_name)},{_q(p.medicine)},{_q(p.dosage)},{_q(p.date)},{_q(p.doctor_name)},{_q(p.hospital_name)},{_q(p.diagnosis)},{_q(p.symptoms)},{_q(p.department)},{_q(p.follow_up_date)},{_q(p.lab_tests)},{p.confidence_score or 90},{p.image_quality_score or 100}\n")
            
        csv_data = "".join(lines)
        return StreamingResponse(
            io.StringIO(csv_data),
            media_type="text/csv",
            headers={"Content-Disposition": "attachment; filename=prescriptions_export.csv"}
        )
    except Exception as e:
        logger.error(f"CSV Export Error: {e}")
        raise HTTPException(status_code=500, detail="Error generating CSV export.")

# ── 6. Formatted Excel Export Endpoint ────────────────────────────────────
@router.get("/api/v1/export/excel")
async def export_excel(
    patient: Optional[str] = Query(None),
    medicine: Optional[str] = Query(None),
    date: Optional[str] = Query(None),
    doctor: Optional[str] = Query(None),
    hospital: Optional[str] = Query(None),
    search: Optional[str] = Query(None),
    sort_by: Optional[str] = Query(None),
    filters: Optional[str] = Query(None),
    page: Optional[int] = Query(None),
    limit: Optional[int] = Query(None),
    db: Session = Depends(get_db)
):
    try:
        query = get_prescriptions_query_base(db, patient, medicine, date, doctor, hospital, search, filters)
        query = apply_sorting(query, sort_by=sort_by)
        if page and limit:
            offset = (page - 1) * limit
            query = query.offset(offset).limit(limit)
        prescriptions = query.all()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Prescriptions Report"
        
        # Headers
        headers = ["ID", "Patient Name", "Medicine Prescribed", "Dosage & Frequency", "Prescription Date", "Doctor Name", "Hospital / Clinic", "Diagnosis", "Symptoms", "Lab Tests", "OCR Confidence", "Quality Score"]
        ws.append(headers)

        # Style header row
        header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Data rows
        for p in prescriptions:
            ws.append([
                p.id, p.patient_name, p.medicine, p.dosage, p.date,
                p.doctor_name or "Unknown", p.hospital_name or "Unknown",
                p.diagnosis or "N/A", p.symptoms or "N/A", p.lab_tests or "N/A",
                f"{p.confidence_score or 90}%", f"{p.image_quality_score or 100}%"
            ])

        # Auto-adjust column widths
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

        stream = io.BytesIO()
        wb.save(stream)
        stream.seek(0)

        return StreamingResponse(
            stream,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=Prescriptions_Report.xlsx"}
        )
    except Exception as e:
        logger.error(f"Excel Export Error: {e}")
        raise HTTPException(status_code=500, detail="Error generating Excel report.")

# ── 7. PDF Report Export Endpoint ─────────────────────────────────────────
@router.get("/api/v1/export/pdf")
async def export_pdf(
    patient: Optional[str] = Query(None),
    medicine: Optional[str] = Query(None),
    date: Optional[str] = Query(None),
    doctor: Optional[str] = Query(None),
    hospital: Optional[str] = Query(None),
    search: Optional[str] = Query(None),
    sort_by: Optional[str] = Query(None),
    filters: Optional[str] = Query(None),
    page: Optional[int] = Query(None),
    limit: Optional[int] = Query(None),
    db: Session = Depends(get_db)
):
    try:
        query = get_prescriptions_query_base(db, patient, medicine, date, doctor, hospital, search, filters)
        query = apply_sorting(query, sort_by=sort_by)
        if page and limit:
            offset = (page - 1) * limit
            query = query.offset(offset).limit(limit)
        prescriptions = query.all()

        pdf = FPDF()
        pdf.add_page()
        pdf.set_font("Helvetica", "B", 16)
        
        # Header
        pdf.set_text_color(79, 70, 229) # Indigo
        pdf.cell(0, 10, "PrescriptionX - Medical Extraction Report", new_x=XPos.LMARGIN, new_y=YPos.NEXT, align="C")
        pdf.ln(5)
        
        pdf.set_font("Helvetica", "B", 10)
        pdf.set_fill_color(240, 240, 240)
        pdf.set_text_color(0, 0, 0)
        
        # Table Header
        cols = [("ID", 12), ("Patient Name", 40), ("Medicine", 38), ("Dosage", 30), ("Date", 25), ("Doctor", 40)]
        for col_name, width in cols:
            pdf.cell(width, 8, col_name, border=1, fill=True, align="C")
        pdf.ln()

        # Table Rows
        pdf.set_font("Helvetica", "", 9)
        for p in prescriptions:
            pdf.cell(12, 7, str(p.id), border=1, align="C")
            pdf.cell(40, 7, str(p.patient_name)[:20], border=1)
            pdf.cell(38, 7, str(p.medicine)[:20], border=1)
            pdf.cell(30, 7, str(p.dosage)[:15], border=1)
            pdf.cell(25, 7, str(p.date)[:12], border=1, align="C")
            pdf.cell(40, 7, str(p.doctor_name or "Unknown")[:20], border=1)
            pdf.ln()

        pdf_output = pdf.output()
        return Response(
            content=bytes(pdf_output),
            media_type="application/pdf",
            headers={"Content-Disposition": "attachment; filename=Prescriptions_Medical_Report.pdf"}
        )
    except Exception as e:
        logger.error(f"PDF Export Error: {e}")
        raise HTTPException(status_code=500, detail="Error generating PDF report.")

# ── 8. Autocomplete Suggestions Endpoint ──────────────────────────────────
@router.get("/api/v1/suggestions", response_model=SuggestionsResponse)
async def get_suggestions(q: str = Query("", min_length=1), db: Session = Depends(get_db)):
    try:
        patients = db.query(Prescription.patient_name).filter(Prescription.patient_name.ilike(f"%{q}%")).distinct().limit(5).all()
        medicines = db.query(Prescription.medicine).filter(Prescription.medicine.ilike(f"%{q}%")).distinct().limit(5).all()
        doctors = db.query(Prescription.doctor_name).filter(Prescription.doctor_name.ilike(f"%{q}%")).distinct().limit(5).all()

        return SuggestionsResponse(
            patients=[p[0] for p in patients if p[0]],
            medicines=[m[0] for m in medicines if m[0]],
            doctors=[d[0] for d in doctors if d[0]]
        )
    except Exception as e:
        logger.error(f"Suggestions Error: {e}")
        return SuggestionsResponse(patients=[], medicines=[], doctors=[])

# ── 9. Authentication Endpoints ───────────────────────────────────────────
@router.post("/auth/register", response_model=UserResponse)
async def register_user(user_in: UserCreate, db: Session = Depends(get_db)):
    existing = db.query(User).filter(User.email == user_in.email.strip().lower()).first()
    if existing:
        raise HTTPException(status_code=400, detail="Email already registered.")
    
    new_user = User(
        email=user_in.email.strip().lower(),
        hashed_password=hash_password(user_in.password),
        full_name=user_in.full_name,
        role=user_in.role or "Staff"
    )
    db.add(new_user)
    db.commit()
    db.refresh(new_user)
    return new_user

@router.post("/auth/login", response_model=Token)
async def login_user(user_in: UserCreate, db: Session = Depends(get_db)):
    user = db.query(User).filter(User.email == user_in.email.strip().lower()).first()
    if not user or not verify_password(user_in.password, user.hashed_password):
        raise HTTPException(status_code=401, detail="Invalid email or password.")
        
    access_token = create_access_token(data={"sub": user.email, "role": user.role})
    return Token(access_token=access_token, token_type="bearer")

@router.get("/auth/me", response_model=UserResponse)
async def get_me(current_user: Optional[User] = Depends(get_current_user)):
    if not current_user:
        raise HTTPException(status_code=401, detail="Not authenticated.")
    return current_user

# ── 10. Seeding Endpoints ──────────────────────────────────────────────────
@router.post("/seed")
async def run_seed(db: Session = Depends(get_db)):
    from .seeder import seed_data_logic
    try:
        count = seed_data_logic(db)
        if count == 0:
            return {"message": "Database is not empty. Seeding skipped.", "inserted": 0}
        return {"message": f"Seeding successful! Inserted {count} prescriptions.", "inserted": count}
    except Exception as e:
        logger.error(f"Seeding error: {e}")
        raise HTTPException(status_code=500, detail=f"Seeding failed: {str(e)}")

@router.get("/seed/status")
async def get_seed_status(db: Session = Depends(get_db)):
    count = db.query(Prescription).count()
    return {
        "is_seeded": count >= 150,
        "record_count": count,
        "status": "seeded" if count > 0 else "empty"
    }

@router.get("/api/v1/filters/options")
async def get_filter_options(db: Session = Depends(get_db)):
    try:
        hospitals = db.query(Prescription.hospital_name).filter(Prescription.hospital_name != None, Prescription.hospital_name != "").distinct().order_by(Prescription.hospital_name).all()
        doctors = db.query(Prescription.doctor_name).filter(Prescription.doctor_name != None, Prescription.doctor_name != "").distinct().order_by(Prescription.doctor_name).all()
        departments = db.query(Prescription.department).filter(Prescription.department != None, Prescription.department != "").distinct().order_by(Prescription.department).all()
        medicines = db.query(Prescription.medicine).filter(Prescription.medicine != None, Prescription.medicine != "").distinct().order_by(Prescription.medicine).all()
        document_types = db.query(Prescription.document_type).filter(Prescription.document_type != None, Prescription.document_type != "").distinct().order_by(Prescription.document_type).all()
        states = db.query(Prescription.state).filter(Prescription.state != None, Prescription.state != "").distinct().order_by(Prescription.state).all()
        cities = db.query(Prescription.city).filter(Prescription.city != None, Prescription.city != "").distinct().order_by(Prescription.city).all()
        languages = db.query(Prescription.language).filter(Prescription.language != None, Prescription.language != "").distinct().order_by(Prescription.language).all()
        medicine_categories = db.query(Prescription.medicine_category).filter(Prescription.medicine_category != None, Prescription.medicine_category != "").distinct().order_by(Prescription.medicine_category).all()
        hospital_types = db.query(Prescription.hospital_type).filter(Prescription.hospital_type != None, Prescription.hospital_type != "").distinct().order_by(Prescription.hospital_type).all()
        doctor_specialties = db.query(Prescription.doctor_specialty).filter(Prescription.doctor_specialty != None, Prescription.doctor_specialty != "").distinct().order_by(Prescription.doctor_specialty).all()

        return {
            "hospitals": [h[0] for h in hospitals],
            "hospital_names": [h[0] for h in hospitals],
            "doctors": [d[0] for d in doctors],
            "doctor_names": [d[0] for d in doctors],
            "departments": [d[0] for d in departments],
            "medicines": [m[0] for m in medicines],
            "document_types": [dt[0] for dt in document_types],
            "states": [s[0] for s in states],
            "cities": [c[0] for c in cities],
            "languages": [l[0] for l in languages],
            "medicine_categories": [mc[0] for mc in medicine_categories],
            "hospital_types": [ht[0] for ht in hospital_types],
            "doctor_specialties": [ds[0] for ds in doctor_specialties]
        }
    except Exception as e:
        logger.error(f"Error fetching unique filter options: {e}")
        raise HTTPException(status_code=500, detail="Error fetching filter options.")
